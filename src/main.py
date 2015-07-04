# The Facebook Online Friend Tracker
# Author: Baraa Hamodi

# Imports
import argparse
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import StaleElementReferenceException
from openpyxl import load_workbook
from os.path import exists
from datetime import datetime
import time

def main():
  # Params
  parser = argparse.ArgumentParser()
  parser.add_argument('--user', dest='facebook_username', required=True, help='the email that you login to Facebook with')
  parser.add_argument('--pass', dest='facebook_password', required=True, help='the password that you login to Facebook with')
  parser.add_argument('--path', dest='path_to_excel_file', required=True, help='the path to the excel file')

  args = parser.parse_args()
  facebook_username = args.facebook_username
  facebook_password = args.facebook_password
  path_to_excel_file = args.path_to_excel_file

  # Verify that the Excel file exists before scraping
  if not exists(path_to_excel_file):
    raise Exception(path_to_excel_file + ' does not exist. Please try again.')

  # Initialize Chrome WebDriver and change default timeout
  driver = webdriver.Chrome()
  driver.implicitly_wait(180)

  # Go to facebook.com and log in with provided credentials
  print('Logging into Facebook...')
  driver.get('https://www.facebook.com/')
  emailBox = driver.find_element_by_id('email')
  emailBox.send_keys(facebook_username)
  passwordBox = driver.find_element_by_id('pass')
  passwordBox.send_keys(facebook_password)
  passwordBox.send_keys(Keys.RETURN)

  # Wait for Facebook to update the number of friends dynamically
  print('Waiting for Facebook to update friends list...')
  time.sleep(120)

  # Try to scrape the value, if it's stale, reload page and try again
  try:
    onlineFriendsCount = int(driver.find_element_by_xpath('//*[@id="fbDockChatBuddylistNub"]/a/span[2]/span').text.strip('()'))
  except StaleElementReferenceException:
    print('StaleElementReferenceException. Retrying...')
    driver.get('https://www.facebook.com/')
    print('Waiting for Facebook to update friends list...')
    time.sleep(120)
    onlineFriendsCount = int(driver.find_element_by_xpath('//*[@id="fbDockChatBuddylistNub"]/a/span[2]/span').text.strip('()'))

  # Close Chrome WebDriver
  driver.close()

  # Get current time
  today = datetime.now().strftime('%Y/%m/%d %H:%M:%S')

  # Load existing spreadsheet
  workbook = load_workbook(path_to_excel_file)
  worksheet = workbook.get_sheet_by_name('Sheet1')

  # Find first empty cell and populate
  for row in worksheet.iter_rows():
    if row[0].value == None:
      row[0].value = today
      row[1].value = onlineFriendsCount
      print('Done! Detected ' + str(onlineFriendsCount) + ' online friends.')
      print('Added: ' + today + ' -> ' + str(onlineFriendsCount) + ' to the spreadsheet.')
      break

  # Save updated spreadsheet
  workbook.save(path_to_excel_file)
